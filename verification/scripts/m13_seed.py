#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
M1-3 真实数据全量导入 · 种子数据生成
从老 Excel 造:product(商品档案)/ machine(2 台真实设备)/ sku_alias(配比底稿 81 条映射)
输出 verification/scripts/out/m13_seed.sql,用 docker exec 灌进 vend_test_imports。

口径:
- 商品档案按 商品编码 去重取首行(同名多行=不同批次进价;一码多名取首名,老账硬伤 M1-6 期初清洗再拆);
- 配比底稿里出现、档案缺失的编码(如 SP068)补建最小档案;
- 别名种子:每个销售名一行,alias_code=SEED-xxx(合成,防 uk 撞),alias_barcode=明细里该销售名最常见条码,
  alias_name=销售名 → 导入通道"条码为主/名称兜底"两条路都可命中。
"""
import re
import sys
from collections import Counter, OrderedDict
from pathlib import Path

from openpyxl import load_workbook

BASE = Path('/Users/yh-1/系统开发/智慧园区')
XLSX = BASE / '小卖铺数据表8.4/小卖铺数据表8.4/自助售卖机业务进销存套表(更新8.1).xlsx'
OUT = Path(__file__).parent / 'out'
OUT.mkdir(exist_ok=True)

wb = load_workbook(XLSX, read_only=True, data_only=True)


def esc(s):
    return str(s).replace('\\', '\\\\').replace("'", "\\'")


# ---------- 1) 商品档案 → product ----------
ws = wb['商品档案']
rows = list(ws.iter_rows(values_only=True))
products = OrderedDict()  # code -> (name, category, unit, ref_cost)
for r in rows[1:]:
    name, code = r[0], r[1]
    if not code or not name:
        continue
    code = str(code).strip()
    if not re.match(r'^SP\d+', code):
        continue
    if code in products:
        continue  # 同码多行取首行
    category = str(r[2]).strip() if r[2] else None
    unit = str(r[5]).strip() if r[5] else '件'
    ref_cost = r[6] if isinstance(r[6], (int, float)) else None
    products[code] = [str(name).strip(), category, unit, ref_cost]

# ---------- 2) 配比底稿 → 销售名映射(+补缺失编码) ----------
ws = wb['配比采购销售编码底稿']
rows = list(ws.iter_rows(values_only=True))
mapping = OrderedDict()  # 销售名 -> SP code
for r in rows[1:]:
    sale_name = r[0]
    code = r[5] or r[7]  # 归集采购商品编码(E/F 对为主,空取 G/H 对——名厨香辣味烤鸭翅根只有第二对)
    if not sale_name or not code:
        continue
    sale_name = str(sale_name).strip()
    code = str(code).strip()
    mapping[sale_name] = code
    if code not in products:
        purchase_name = r[4] or r[6] or sale_name
        products[code] = [str(purchase_name).strip(), None, '件', None]
        print(f'补建缺失档案: {code} {purchase_name}')

# ---------- 3) 销售明细 → 设备 + 每销售名最常见条码 ----------
ws = wb['销售明细表6-7月']
rows = list(ws.iter_rows(values_only=True))
devices = OrderedDict()  # device_id -> 设备名称
barcode_votes = {}  # 销售名 -> Counter(条码)
for r in rows[1:]:
    if not r[9]:
        continue
    name = str(r[0]).strip()
    barcode = str(r[2]).strip() if r[2] else None
    dev_id, dev_name = str(r[5]).strip(), str(r[6]).strip()
    devices.setdefault(dev_id, dev_name)
    if barcode:
        barcode_votes.setdefault(name, Counter())[barcode] += 1

missing = [n for n in barcode_votes if n not in mapping]
unmapped = [n for n in mapping if n not in barcode_votes]
print(f'商品 {len(products)} · 映射 {len(mapping)} · 设备 {len(devices)} · 明细有但映射缺 {missing}')

# ---------- 4) 生成 SQL ----------
sql = ['SET NAMES utf8mb4;']
for code, (name, category, unit, ref_cost) in products.items():
    cols = f"'{esc(code)}', '{esc(name)}', " \
           + (f"'{esc(category)}'" if category else 'NULL') + f", '{esc(unit)}'" \
           + (f", {ref_cost}" if ref_cost is not None else ', NULL')
    sql.append(
        "INSERT INTO yc_vend_product (sku_code, product_name, category, unit, ref_cost, create_user)"
        f" VALUES ({cols}, 1);")
for i, (dev_id, dev_name) in enumerate(devices.items(), 1):
    sql.append(
        "INSERT INTO yc_vend_machine (machine_code, machine_name, device_id, machine_status, create_user)"
        f" VALUES ('VM{i:03d}', '{esc(dev_name)}', '{esc(dev_id)}', '在线', 1);")
for i, (sale_name, code) in enumerate(mapping.items(), 1):
    barcode = ''
    if sale_name in barcode_votes:
        barcode = barcode_votes[sale_name].most_common(1)[0][0]
    sql.append(
        "INSERT INTO yc_vend_sku_alias (alias_code, alias_barcode, alias_name, product_id, bind_source, create_user)"
        f" SELECT 'SEED-{i:03d}', '{esc(barcode)}', '{esc(sale_name)}', id, '人工', 1"
        f" FROM yc_vend_product WHERE sku_code='{esc(code)}';")

out = OUT / 'm13_seed.sql'
out.write_text('\n'.join(sql), encoding='utf-8')
print(f'已生成 {out} · {len(sql)-1} 条 INSERT')
if missing:
    sys.exit('❌ 有销售名缺映射,导入会进待绑定,先核对配比底稿')
