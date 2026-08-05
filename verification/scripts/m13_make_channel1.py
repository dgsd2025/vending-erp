#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
M1-3 真实数据全量导入 · 通道1文件生成
「销售明细表6-7月」sheet(5135 行 fanmaiji 导出)→ 通道1「出货明细」格式 xlsx。
列取导入中心通道1规格(必填:订单号/商品名称/出货数量/商品金额(元)/设备ID/订单类型/出货时间);
出货时间保留原文本(fanmaiji 导出即文本,数据字典脏点①,导入端已兜多格式解析)。
输出 verification/scripts/out/通道1-出货明细-全量.xlsx
"""
from pathlib import Path

from openpyxl import Workbook, load_workbook

BASE = Path('/Users/yh-1/系统开发/智慧园区')
SRC = BASE / '小卖铺数据表8.4/小卖铺数据表8.4/自助售卖机业务进销存套表(更新8.1).xlsx'
OUT = Path(__file__).parent / 'out' / '通道1-出货明细-全量.xlsx'
OUT.parent.mkdir(exist_ok=True)

HEADERS = ['订单号', '商品名称', '商品条形码', '出货数量', '货道号', '设备ID', '设备名称',
           '商品金额(元)', '订单类型', '支付方式', '出货时间']

src = load_workbook(SRC, read_only=True, data_only=True)['销售明细表6-7月']
wb = Workbook(write_only=True)
ws = wb.create_sheet('出货明细')
ws.append(HEADERS)

count, total = 0, 0.0
for r in src.iter_rows(min_row=2, values_only=True):
    # 源列:0商品名称 1分类 2条形码 3出货数量 4货道号 5设备ID 6设备名称 7运营商 8商品金额 9订单号 10订单类型 11支付方式 12出货时间
    if not r[9]:
        continue
    ws.append([
        str(r[9]).strip(),                       # 订单号
        str(r[0]).strip(),                       # 商品名称
        str(r[2]).strip() if r[2] else None,     # 商品条形码
        r[3],                                    # 出货数量
        str(r[4]) if r[4] is not None else None, # 货道号
        str(r[5]).strip(),                       # 设备ID
        str(r[6]).strip() if r[6] else None,     # 设备名称
        r[8],                                    # 商品金额(元)
        str(r[10]).strip(),                      # 订单类型(正常订单)
        str(r[11]).strip() if r[11] else None,   # 支付方式
        str(r[12]).strip(),                      # 出货时间(文本)
    ])
    count += 1
    total += float(r[8] or 0)

wb.save(OUT)
print(f'已生成 {OUT} · {count} 行 · Σ商品金额 = {total:.2f}(冲刺0基准应为 25113.50)')
